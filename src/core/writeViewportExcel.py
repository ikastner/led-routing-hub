#!/usr/bin/env python3
"""Génère un Excel viewport (N bandes RGB + devices) pour le profil 32×32."""

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


def main() -> None:
    payload_path = Path(sys.argv[1])
    data = json.loads(payload_path.read_text(encoding="utf-8"))
    out = Path(data["outPath"])
    out.parent.mkdir(parents=True, exist_ok=True)

    source = Path(data.get("sourceXlsx") or "")
    col_start = int(data["colStart"])
    col_end = int(data["colEnd"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # En-tête (ligne 1) — repris de la source si possible
    headers = ["Band", "EntityStart", "EntityEnd", "IP", "Universe"]
    if source.is_file():
        src = load_workbook(source, read_only=True, data_only=True)
        sheet = src.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            headers = [c if c is not None else headers[i] for i, c in enumerate(header_row[:5])]
        # Devices : lignes ≥ 130 sur le mapping Glassworks 128
        device_rows = []
        for row in sheet.iter_rows(min_row=130, values_only=True):
            if row and row[0]:
                device_rows.append(list(row[:5]))
        src.close()
    else:
        device_rows = []

    ws.append(headers)

    for row in data["rows"]:
        ws.append(
            [
                row["name"],
                row["entityStart"],
                row["entityEnd"],
                row["controllerIp"],
                row["universe"],
            ]
        )

    # Devices à la suite des bandes (parseEcran détecte projecteur/lyre par nom)
    if device_rows:
        for dr in device_rows:
            ws.append(dr)
    else:
        for d in data.get("devices") or []:
            if d.get("type") == "rgbw":
                ws.append(
                    [
                        d.get("name") or "Projector",
                        1,
                        4,
                        d.get("controllerIp") or "192.168.1.48",
                        d.get("universe") or 33,
                    ]
                )
            else:
                ws.append(
                    [
                        d.get("name") or "Lyre",
                        d.get("dmxChannelStart"),
                        d.get("dmxChannelEnd"),
                        d.get("controllerIp") or "192.168.1.48",
                        d.get("universe") or 33,
                    ]
                )

    wb.save(out)
    print(f"Excel viewport écrit : {out} ({col_end - col_start} bandes)")


if __name__ == "__main__":
    main()
